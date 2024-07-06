import uvicorn
import asyncio
import csv
import math

from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from starlette.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from io import StringIO, BytesIO
from kpi import main_calculations

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Analyzing AS-IS & TO-BE
@app.post("/upload/")
async def upload_file(file: UploadFile = File(...)):
    try:
        contents = await file.read()

        result = main_calculations(contents)

        return JSONResponse(status_code=200, content={"analyze": result})
    except Exception as e:
        print(str(e))
        return JSONResponse(status_code=500, content={"message": "An error occurred", "details": str(e)})


def custom_round_up(number):
    try:
        if number > 0:
            return int(number) + 1 if (number - int(number)) >= 0.5 else int(number)
        elif number < 0:
            return int(number) if (number - int(number)) > -0.5 else int(number) - 1
        else:
            return 0
    except Exception as e:
        print(e)
        return 0


def normal_round(n, decimals=0):
    try:
        expoN = n * 10 ** decimals
        if abs(expoN) - abs(math.floor(expoN)) < 0.5:
            return math.floor(expoN) / 10 ** decimals
        return math.ceil(expoN) / 10 ** decimals
    except Exception as e:
        print(e)
        return 0

# Overall Analyzing
@app.post("/overall/")
async def overallAnalyze(result: dict):
    try:
        data = result

        as_is_dict = data['result']['as_is']
        to_be_dict = data['result']['to_be']

        as_is_complexity = normal_round(calculate_complexity(as_is_dict), 2)
        to_be_complexity = normal_round(calculate_complexity(to_be_dict), 2)
        effect = calculate_effect(as_is_dict, to_be_dict)
        improvement = calculate_improvement(as_is_complexity, to_be_complexity)

        return JSONResponse(status_code=200, content={"as_is_complexity": as_is_complexity, "to_be_complexity": to_be_complexity, "effect": effect, "improvement": improvement})
    except Exception as e:
        print(str(e))
        return JSONResponse(status_code=500, content={"message": "An error occurred", "details": str(e)})
    

def calculate_complexity(counts):
    try:
        complexity_weights = {
            'CC': 0.25,
            'CO': 0.05,
            'CS': 0.05,
            'DC': 0.20,
            'DP': 0.07,
            'S': 0.10,
            'R': 0.06,
            'M': 0.15,
            'T': 0.04,
            'Q': 0.03
        }
        complexity_score = sum(counts[item] * complexity_weights[item] for item in counts)
        return complexity_score
    except Exception as e:
        print(e)

def calculate_effect(as_is_dict, to_be_dict):
    effect_dict = {}
    for key in as_is_dict:
        try:
            if key in to_be_dict:
                if as_is_dict[key] != 0:
                    effect = round(((to_be_dict[key] - as_is_dict[key]) / as_is_dict[key]) * 100, 2)
                else:
                    effect = 100 if to_be_dict[key] != 0 else 0
                effect_dict[key] = -1 * custom_round_up(effect)
            else:
                effect_dict[key] = 0
        except ZeroDivisionError:
            effect_dict[key] = 0
        except Exception as e:
            effect_dict[key] = 0
            print(f"An unexpected error occurred for key {key}: {e}")

    return effect_dict


def calculate_improvement(as_is_complexity, to_be_complexity):
    try:
        if to_be_complexity == 0:
            print("Error: TO_BE complexity is zero, cannot calculate improvement.")
            return 0
        return custom_round_up(((to_be_complexity - as_is_complexity) / as_is_complexity) * 100)
    except Exception as e:
        print(e)
        return 0


# Download Overall .CSV
@app.post('/download-overall/csv')
async def download_csv(result: dict):
    try:
        data = result
        
        as_is_dict = data['result']['as_is']
        to_be_dict = data['result']['to_be']

        effect_dict = data['result']['overall']['effect']
        as_is_complexity = data['result']['overall']['as_is_complexity']
        to_be_complexity = data['result']['overall']['to_be_complexity']
        improvement = data['result']['overall']['improvement']

        data = [
            ["Код и параметр оптимизации", "Показатель как есть", "Показатель как будет", "Эффект"],
            ["CC: Неспосредственные контакты", as_is_dict['CC'], to_be_dict['CC'], f"{effect_dict['CC']}%"],
            ["CO: Опосредованные контакты", as_is_dict['CO'], to_be_dict['CO'], f"{effect_dict['CO']}%"],
            ["CS: Контакты с подрядчиками", as_is_dict['CS'], to_be_dict['CS'], f"{effect_dict['CS']}%"],
            ["DC: Входящие документы", as_is_dict['DC'], to_be_dict['DC'], f"{effect_dict['DC']}%"],
            ["DP: Порождаемые документы", as_is_dict['DP'], to_be_dict['DP'], f"{effect_dict['DP']}%"],
            ["S: Шаги", as_is_dict['S'], to_be_dict['S'], f"{effect_dict['S']}%"],
            ["R: Бизнес-роли", as_is_dict['R'], to_be_dict['R'], f"{effect_dict['R']}%"],
            ["M: Руководители", as_is_dict['M'], to_be_dict['M'], f"{effect_dict['M']}%"],
            ["T: Передачи", as_is_dict['T'], to_be_dict['T'], f"{effect_dict['T']}%"],
            ["Q: Перемещения", as_is_dict['Q'], to_be_dict['Q'], f"{effect_dict['Q']}%"],
            ["Сложность процесса", as_is_complexity, to_be_complexity],
            ["Относительная степень улучшения", f"{improvement}%"]
        ]

        output = StringIO()
        writer = csv.writer(output)
        writer.writerows(data)
        output.seek(0)

        headers = {
            "Content-Disposition": "attachment; filename=data.csv",
            "Content-Type": "text/csv; charset=utf-8"
        }

        return Response(content=output.getvalue(), media_type="text/csv", headers=headers)
    
    except Exception as e:
        print(e)

# Download Overall .XLSX
@app.post("/download-overall/xlsx")
async def download_xlsx(result: dict):
    try:
        data = result
        
        as_is_dict = data['result']['as_is']
        to_be_dict = data['result']['to_be']

        effect_dict = data['result']['overall']['effect']
        as_is_complexity = str(data['result']['overall']['as_is_complexity']).replace('.', ',')
        to_be_complexity = str(data['result']['overall']['to_be_complexity']).replace('.', ',')
        improvement = data['result']['overall']['improvement']

        data = [
            ["Код и параметр оптимизации", "Показатель как есть", "Показатель как будет", "Эффект"],
            ["CC: Неспосредственные контакты", as_is_dict['CC'], to_be_dict['CC'], f"{effect_dict['CC']}%"],
            ["CO: Опосредованные контакты", as_is_dict['CO'], to_be_dict['CO'], f"{effect_dict['CO']}%"],
            ["CS: Контакты с подрядчиками", as_is_dict['CS'], to_be_dict['CS'], f"{effect_dict['CS']}%"],
            ["DC: Входящие документы", as_is_dict['DC'], to_be_dict['DC'], f"{effect_dict['DC']}%"],
            ["DP: Порождаемые документы", as_is_dict['DP'], to_be_dict['DP'], f"{effect_dict['DP']}%"],
            ["S: Шаги", as_is_dict['S'], to_be_dict['S'], f"{effect_dict['S']}%"],
            ["R: Бизнес-роли", as_is_dict['R'], to_be_dict['R'], f"{effect_dict['R']}%"],
            ["M: Руководители", as_is_dict['M'], to_be_dict['M'], f"{effect_dict['M']}%"],
            ["T: Передачи", as_is_dict['T'], to_be_dict['T'], f"{effect_dict['T']}%"],
            ["Q: Перемещения", as_is_dict['Q'], to_be_dict['Q'], f"{effect_dict['Q']}%"],
            ["Сложность процесса", f"{as_is_complexity}", f"{to_be_complexity}", ""],
            ["Относительная степень улучшения", "", "", f"{improvement}%"]
        ]

        wb = Workbook()
        ws = wb.active

        header_font = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border_style = Side(style='thin', color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        bold_font = Font(name="Times New Roman", size=12, bold=True)
        bold_italic_white_font = Font(name="Times New Roman", size=12, bold=True, italic=True, color="FFFFFF")
        green_fill = PatternFill(start_color="38A454", end_color="38A454", fill_type="solid")
        regular_font = Font(name="Times New Roman", size=12)

        # Apply styles to header row
        for col_num, header in enumerate(data[0], start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Write the data to the worksheet with styles
        for row_num, row_data in enumerate(data[1:], start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.font = regular_font
                cell.border = border
                cell.alignment = center_alignment if col_num > 1 else left_alignment
                if row_num == 12:  # Apply bold font for Сложность процесса (Ср)
                    cell.font = bold_font
                if row_num == 13:  # Apply green fill and bold italic white font for Относительная степень улучшений
                    cell.font = bold_italic_white_font
                    cell.fill = green_fill

        # Set column widths for better readability
        column_widths = [35, 20, 20, 10]
        for i, column_width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = column_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        headers = {
            "Content-Disposition": "attachment; filename=data.xlsx"
        }

        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    
    except Exception as e:
        print(e)


# Server Settings & Main
async def main():
    try:
        config = uvicorn.Config("server:app", host = '0.0.0.0', port = 3000, log_level = 'info')
        server = uvicorn.Server(config = config)
        await server.serve()
    except Exception as e:
        print(e)

if __name__ == '__main__':
    try:
        asyncio.run(main())
    except Exception as e:
        print(e)